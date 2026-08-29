#!/usr/bin/env python3
"""
Populate assets/js/media.js from a spreadsheet export of lesson video URLs.

    python3 tools/import-media.py lessons.csv
    python3 tools/import-media.py WPI_LMS_Migration_Manifest.xlsx --dry-run

A .xlsx is read from its "Lessons" sheet and understands the migration
manifest's own columns (Vimeo ID, Vimeo Hash, Video, Type, Duration (s)).

The CSV needs a course column, a lesson column and a URL column. Header names
are matched loosely, so all of these work:

    course, lesson, url
    Course Name, Lesson Title, Video URL
    course_title, lesson_title, src

Optional columns, same loose matching: poster, captions, duration, download.

Lessons are matched to the catalog in assets/js/data.js by normalised title
(case, punctuation and whitespace are ignored), so an export does not have to
match the catalog character for character. Every row that cannot be matched is
reported at the end with the closest candidates — nothing is silently dropped.

Rows are written keyed by lesson title rather than index, so the map keeps
working if a course is reordered later.
"""

import argparse
import csv
import json
import os
import re
import sys
from difflib import get_close_matches

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, '..'))
DATA_JS = os.path.join(ROOT, 'assets', 'js', 'data.js')
MEDIA_JS = os.path.join(ROOT, 'assets', 'js', 'media.js')

COLUMNS = {
    'course':   ('course', 'course_title', 'course name', 'coursename', 'programme', 'program'),
    'lesson':   ('lesson', 'lesson_title', 'lesson name', 'lessonname', 'title', 'module'),
    'src':      ('url', 'src', 'video url', 'video_url', 'videourl', 'link', 'source', 'target_path', 'target filename'),
    'vimeo_id': ('vimeo_id', 'vimeo id', 'vimeoid'),
    'vimeo_hash': ('vimeo_hash', 'vimeo hash', 'vimeohash', 'hash'),
    'has_video': ('video',),
    'poster':   ('poster', 'thumbnail', 'thumb', 'image'),
    'captions': ('captions', 'caption', 'vtt', 'subtitles', 'subtitle'),
    'duration': ('duration s', 'duration sec', 'duration_sec', 'seconds',
                 'duration', 'length', 'runtime'),
    'download': ('download', 'file', 'attachment', 'resource'),
}


def norm(value):
    """Fold case, punctuation and whitespace so titles compare loosely."""
    value = (value or '').lower().replace('&', ' and ')
    value = re.sub(r'[‘’“”]', '', value)
    value = re.sub(r'[^a-z0-9]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def read_xlsx(path):
    """Rows from the migration manifest's Lessons sheet."""
    try:
        import openpyxl
    except ImportError:
        sys.exit('Reading .xlsx needs openpyxl. Run:  pip install openpyxl')
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book['Lessons'] if 'Lessons' in book.sheetnames else book[book.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = ['' if h is None else str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        out.append({header[i]: ('' if v is None else str(v).strip())
                    for i, v in enumerate(row) if i < len(header)})
    return out, header


def load_catalog():
    """Read the course list straight out of assets/js/data.js."""
    with open(DATA_JS, encoding='utf-8') as fh:
        source = fh.read()
    payload = json.loads(source[source.index('{'):source.rindex(';')])
    return payload['courses']


def map_columns(fieldnames):
    """Match the export's headers to the fields we understand."""
    found, seen = {}, {}
    for name in fieldnames or []:
        seen[norm(name)] = name
    for key, options in COLUMNS.items():
        for option in options:
            if norm(option) in seen:
                found[key] = seen[norm(option)]
                break
    return found


def detect_type(url):
    """Best guess at how a URL should be played, mirroring the player."""
    low = (url or '').lower()
    if 'vimeo.com' in low:
        return 'vimeo'
    if 'youtube.com' in low or 'youtu.be' in low:
        return 'youtube'
    if 'wistia' in low:
        return 'wistia'
    if '.m3u8' in low:
        return 'hls'
    if '.webm' in low:
        return 'webm'
    if '.mp4' in low or '.m4v' in low or '.mov' in low:
        return 'mp4'
    return None


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('csv_path', help='CSV export of lesson video URLs')
    parser.add_argument('--dry-run', action='store_true',
                        help='report what would be written without writing it')
    parser.add_argument('--base', default='',
                        help='prefix for relative paths, e.g. a CDN URL or assets/media '
                             '— use with a file-path column to self-host instead of embedding')
    parser.add_argument('--fuzzy', action='store_true',
                        help='accept close title matches, printing each one for review')
    args = parser.parse_args()

    if not os.path.isfile(args.csv_path):
        sys.exit('No such file: %s' % args.csv_path)

    courses = load_catalog()
    by_course = {}
    for course in courses:
        by_course[norm(course['title'])] = course
        by_course[norm(course['slug'])] = course

    if args.csv_path.lower().endswith(('.xlsx', '.xlsm')):
        rows, fieldnames = read_xlsx(args.csv_path)
    else:
        with open(args.csv_path, newline='', encoding='utf-8-sig') as fh:
            reader = csv.DictReader(fh)
            fieldnames = reader.fieldnames
            rows = list(reader)

    cols = map_columns(fieldnames)
    missing = [k for k in ('course', 'lesson') if k not in cols]
    if missing:
        sys.exit('Missing a %s column. Headers seen: %s'
                 % (' and a '.join(missing), ', '.join(fieldnames or [])))
    if 'src' not in cols and 'vimeo_id' not in cols:
        sys.exit('No URL column and no Vimeo ID column. Headers seen: %s'
                 % ', '.join(fieldnames or []))

    media = {}
    matched = 0
    skipped_no_video = 0
    fuzzy_used = []
    unmatched = []

    for n, row in enumerate(rows, start=2):
        course_raw = (row.get(cols['course']) or '').strip()
        lesson_raw = (row.get(cols['lesson']) or '').strip()

        # Rows the manifest marks as having no video (quizzes, documents)
        # are skipped rather than counted as failures to match.
        if 'has_video' in cols:
            flag = (row.get(cols['has_video']) or '').strip().lower()
            if flag in ('no', 'false', '0'):
                skipped_no_video += 1
                continue

        vimeo_id = (row.get(cols.get('vimeo_id', '')) or '').strip()
        vimeo_hash = (row.get(cols.get('vimeo_hash', '')) or '').strip()
        src = (row.get(cols.get('src', '')) or '').strip()

        if args.base and src:
            src = args.base.rstrip('/') + '/' + src.lstrip('/')
        elif vimeo_id:
            # An unlisted video needs its hash; keep them as separate fields so
            # the player can build the embed URL either way.
            src = 'https://vimeo.com/' + vimeo_id
        if not src:
            continue

        course = by_course.get(norm(course_raw))
        if not course:
            near = get_close_matches(norm(course_raw), list(by_course), n=1, cutoff=0.6)
            unmatched.append((n, course_raw, lesson_raw, 'no such course',
                              by_course[near[0]]['title'] if near else ''))
            continue

        titles = {norm(l['t']): l['t'] for l in course['lessons']}
        title = titles.get(norm(lesson_raw))
        if not title:
            near = get_close_matches(norm(lesson_raw), list(titles), n=1, cutoff=0.6)
            if near and args.fuzzy:
                title = titles[near[0]]
                fuzzy_used.append((lesson_raw, title))
            else:
                unmatched.append((n, course_raw, lesson_raw, 'no such lesson in that course',
                                  titles[near[0]] if near else ''))
                continue

        entry = {'src': src}
        kind = detect_type(src)
        if kind:
            entry['type'] = kind
        if vimeo_hash and entry.get('type') == 'vimeo':
            entry['hash'] = vimeo_hash
        for field in ('poster', 'captions', 'download'):
            if field in cols:
                value = (row.get(cols[field]) or '').strip()
                if value:
                    entry[field] = value
        if 'duration' in cols:
            value = (row.get(cols['duration']) or '').strip()
            if value:
                seconds = None
                if ':' in value:
                    # h:mm:ss or mm:ss, as a formatted duration column carries it
                    try:
                        parts = [float(x) for x in value.split(':')]
                        seconds = 0
                        for part in parts:
                            seconds = seconds * 60 + part
                    except ValueError:
                        seconds = None
                else:
                    try:
                        seconds = float(value)
                    except ValueError:
                        seconds = None
                if seconds:
                    entry['duration'] = int(seconds)

        media.setdefault(course['slug'], {})[title] = entry
        matched += 1

    total_lessons = sum(len(c['lessons']) for c in courses)
    print('%d rows read, %d matched to a lesson' % (len(rows), matched))
    if skipped_no_video:
        print('%d rows skipped as having no video (quizzes and documents)' % skipped_no_video)
    if fuzzy_used:
        print('\n%d title%s accepted by close match — check these:'
              % (len(fuzzy_used), '' if len(fuzzy_used) == 1 else 's'))
        for a, b in fuzzy_used:
            print('   %-52s -> %s' % (a[:52], b))
    print('%d of %d lessons in the catalog now have a source'
          % (matched, total_lessons))

    if unmatched:
        print('\n%d row%s could not be matched:'
              % (len(unmatched), '' if len(unmatched) == 1 else 's'))
        for line, course_raw, lesson_raw, why, suggestion in unmatched[:40]:
            hint = '  (closest: %s)' % suggestion if suggestion else ''
            print('  line %-5d %s / %s — %s%s' % (line, course_raw, lesson_raw, why, hint))
        if len(unmatched) > 40:
            print('  ... and %d more' % (len(unmatched) - 40))

    covered = {}
    for course in courses:
        have = len(media.get(course['slug'], {}))
        if have < len(course['lessons']):
            covered[course['title']] = (have, len(course['lessons']))
    if covered:
        print('\nCourses not fully covered:')
        for title, (have, total) in sorted(covered.items()):
            print('  %-52s %d/%d' % (title[:52], have, total))

    if args.dry_run:
        print('\nDry run — %s not written.' % os.path.relpath(MEDIA_JS, ROOT))
        return

    with open(MEDIA_JS, 'w', encoding='utf-8') as fh:
        fh.write('/* Generated by tools/import-media.py from %s.\n'
                 '   Keyed by course slug, then by exact lesson title.\n\n'
                 '   A Vimeo `hash` is the unlisted-video key: anyone holding the id\n'
                 '   and hash can view that video. Keep this repository private, or\n'
                 '   self-host the files and drop the hashes. */\n'
                 % os.path.basename(args.csv_path))
        fh.write('window.WPI_MEDIA = ')
        json.dump(media, fh, ensure_ascii=False, indent=1, sort_keys=True)
        fh.write(';\n')
    print('\nwrote %s' % os.path.relpath(MEDIA_JS, ROOT))


if __name__ == '__main__':
    main()
