#!/usr/bin/env python3
"""
Correct lesson types in assets/js/data.js from the LMS migration manifest.

    python3 tools/sync-lesson-types.py WPI_LMS_Migration_Manifest.xlsx
    python3 tools/sync-lesson-types.py WPI_LMS_Migration_Manifest.xlsx --dry-run

The catalog was built by reading the site archive, which shows how a lesson was
presented but not always what it actually held. The manifest is the platform's
own export and says outright, per lesson, whether there is a video and whether
there is a document. Where the two disagree the manifest wins.

This matters because a lesson typed `video` with no video behind it renders an
empty player stage telling the rep to go and edit a JavaScript file. Three
Customer Connect lessons and the PCI Program lesson do exactly that today, and
none of them ever had a video to play.

Types written:

    quiz      manifest Type is Quiz
    video     manifest says Video = Yes
    download  no video, but Doc = Yes
    text      neither - a written lesson or an outbound link

Titles are matched with case, punctuation and accents folded, so the catalog
does not have to match the manifest character for character.
"""

import argparse
import json
import os
import re
import sys
import unicodedata

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, '..'))
DATA_JS = os.path.join(ROOT, 'assets', 'js', 'data.js')


def norm(value):
    """Fold case, punctuation, accents and whitespace so titles compare loosely."""
    value = unicodedata.normalize('NFKD', str(value or ''))
    value = value.lower().replace('&', ' and ')
    value = re.sub(r'[^a-z0-9]+', ' ', value)
    return ' '.join(value.split())


def kind_of(row):
    if str(row.get('Type') or '').strip().lower() == 'quiz':
        return 'quiz'
    if str(row.get('Video') or '').strip().lower() == 'yes':
        return 'video'
    if str(row.get('Doc') or '').strip().lower() == 'yes':
        return 'download'
    return 'text'


def read_manifest(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl is required to read the manifest: pip install openpyxl')
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book['Lessons'] if 'Lessons' in book.sheetnames else book[book.sheetnames[0]]
    rows = list(sheet.values)
    header = [str(h) if h is not None else '' for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('manifest', help='WPI_LMS_Migration_Manifest.xlsx')
    parser.add_argument('--dry-run', action='store_true',
                        help='report what would change without writing')
    args = parser.parse_args()

    manifest = read_manifest(args.manifest)

    # First spelling wins: a course can repeat a lesson title across sections,
    # and the manifest lists them in the order the platform did.
    truth = {}
    for row in manifest:
        key = (norm(row.get('Course')), norm(row.get('Lesson')))
        truth.setdefault(key, kind_of(row))

    source = open(DATA_JS, encoding='utf-8').read()
    head = source[:source.index('window.WPI = ')]
    payload = source[source.index('window.WPI = ') + len('window.WPI = '):]
    data = json.loads(payload.rstrip().rstrip(';'))

    changes, unmatched = [], []
    for course in data['courses']:
        for lesson in course['lessons']:
            key = (norm(course['title']), norm(lesson['t']))
            want = truth.get(key)
            if want is None:
                unmatched.append((course['title'], lesson['t']))
                continue
            if want != lesson['k']:
                changes.append((course['title'], lesson['t'], lesson['k'], want))
                lesson['k'] = want

    if changes:
        print('%d lesson type(s) corrected from the manifest:' % len(changes))
        for course, lesson, was, now in changes:
            print('  %-46s %-44s %s -> %s' % (course[:46], lesson[:44], was, now))
    else:
        print('every lesson type already agrees with the manifest')

    if unmatched:
        print('\n%d catalog lesson(s) are not in the manifest and were left alone:'
              % len(unmatched))
        for course, lesson in unmatched:
            print('  %-46s %s' % (course[:46], lesson))

    counts = {}
    for course in data['courses']:
        for lesson in course['lessons']:
            counts[lesson['k']] = counts.get(lesson['k'], 0) + 1
    print('\nlesson types now: %s' % ', '.join(
        '%d %s' % (v, k) for k, v in sorted(counts.items(), key=lambda kv: -kv[1])))

    if args.dry_run:
        print('\nDry run — %s not written.' % os.path.relpath(DATA_JS, ROOT))
        return
    if not changes:
        return

    with open(DATA_JS, 'w', encoding='utf-8') as fh:
        fh.write(head)
        fh.write('window.WPI = ')
        json.dump(data, fh, ensure_ascii=False, indent=1)
        fh.write(';\n')
    print('\nwrote %s' % os.path.relpath(DATA_JS, ROOT))


if __name__ == '__main__':
    main()
